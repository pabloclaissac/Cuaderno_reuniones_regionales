
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from PIL import Image, ImageTk
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import tkinter.font as tkfont

# Color burdeos para los botones
BURDEOS = "#800020"
AZUL_OSCURO = "#003366"
GRIS_OSCURO = "#333333"

# Ruta del archivo Excel
EXCEL_FILE = "Planificación 2025.xlsx"
SHEET_NAME = "Hoja3"

class ExcelManager:
    """Clase para manejar operaciones con Excel"""
    @staticmethod
    def get_next_empty_row(worksheet):
        """Encuentra la próxima fila vacía en la hoja de cálculo"""
        row = 1
        while worksheet.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    @staticmethod
    def save_to_excel(tomo_name, tema, detalle):
        """Guarda los datos en el archivo Excel"""
        try:
            # Intentar cargar el archivo existente
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
            except FileNotFoundError:
                # Si el archivo no existe, crear uno nuevo
                wb = openpyxl.Workbook()
                # Eliminar la hoja por defecto si es necesario
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
            
            # Obtener o crear la hoja
            if SHEET_NAME in wb.sheetnames:
                sheet = wb[SHEET_NAME]
            else:
                sheet = wb.create_sheet(SHEET_NAME)
                # Crear encabezados si es una hoja nueva
                sheet['A1'] = "Dirección Regional"
                sheet['B1'] = "Fecha de Reunión"
                sheet['C1'] = "Ítem de monitoreo"
                sheet['D1'] = "Detalle"
            
            # Encontrar la próxima fila vacía
            next_row = ExcelManager.get_next_empty_row(sheet)
            
            # Escribir los datos
            sheet[f'A{next_row}'] = tomo_name
            sheet[f'B{next_row}'] = datetime.now().strftime("%d/%m/%Y")
            sheet[f'C{next_row}'] = tema
            sheet[f'D{next_row}'] = detalle
            
            # Guardar el archivo
            wb.save(EXCEL_FILE)
            return True
        except Exception as e:
            print(f"Error al guardar en Excel: {e}")
            return False


class CommentWindow(tk.Toplevel):
    """Ventana para crear nuevos comentarios"""
    def __init__(self, parent, callback=None):
        super().__init__(parent)
        self.title("Nuevo Comentario")
        self.geometry("400x300")
        self.resizable(True, True)
        self.callback = callback
        
        # Frame principal
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Área de texto para el comentario
        tk.Label(main_frame, text="Contenido del comentario:").pack(anchor=tk.W, pady=(0, 5))
        
        self.text_area = tk.Text(
            main_frame, 
            wrap="word", 
            font=("Arial", 9),
            padx=10,
            pady=10,
            height=8
        )
        self.text_area.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Frame para botones
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        # Botón Guardar
        save_btn = tk.Button(
            btn_frame, 
            text="Guardar", 
            bg=BURDEOS, 
            fg="white",
            width=10,
            command=self.save_comment
        )
        save_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Botón Cancelar
        cancel_btn = tk.Button(
            btn_frame, 
            text="Cancelar", 
            bg="#606060", 
            fg="white",
            width=10,
            command=self.destroy
        )
        cancel_btn.pack(side=tk.RIGHT)
        
        # Configurar enfoque
        self.text_area.focus_set()
    
    def save_comment(self):
        """Guarda el comentario y cierra la ventana"""
        content = self.text_area.get("1.0", "end-1c")
        if self.callback:
            self.callback(content)
        self.destroy()


class EditCommentWindow(tk.Toplevel):
    """Ventana para editar comentarios existentes"""
    def __init__(self, parent, content, callback=None):
        super().__init__(parent)
        self.title("Editar Comentario")
        self.geometry("400x300")
        self.resizable(True, True)
        self.callback = callback
        
        # Frame principal
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Área de texto para el comentario
        tk.Label(main_frame, text="Contenido del comentario:").pack(anchor=tk.W, pady=(0, 5))
        
        self.text_area = tk.Text(
            main_frame, 
            wrap="word", 
            font=("Arial", 9),
            padx=10,
            pady=10,
            height=8
        )
        self.text_area.insert("1.0", content)
        self.text_area.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Frame para botones
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        # Botón Guardar
        save_btn = tk.Button(
            btn_frame, 
            text="Guardar", 
            bg=BURDEOS, 
            fg="white",
            width=10,
            command=self.save_comment
        )
        save_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Botón Cerrar
        close_btn = tk.Button(
            btn_frame, 
            text="Cerrar", 
            bg="#606060", 
            fg="white",
            width=10,
            command=self.destroy
        )
        close_btn.pack(side=tk.RIGHT)
        
        # Configurar enfoque
        self.text_area.focus_set()
    
    def save_comment(self):
        """Guarda el comentario y cierra la ventana"""
        content = self.text_area.get("1.0", "end-1c")
        if self.callback:
            self.callback(content)
        self.destroy()


class CommentSearchWindow(tk.Toplevel):
    """Ventana para buscar y gestionar comentarios"""
    def __init__(self, parent, editor):
        super().__init__(parent)
        self.title("Buscar y Gestionar Comentarios")
        self.geometry("600x400")
        self.resizable(True, True)
        self.editor = editor
        
        # Frame principal
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Frame de búsqueda
        search_frame = tk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(search_frame, text="Buscar en comentarios:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_entry = tk.Entry(search_frame, width=30)
        self.search_entry.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.search_entry.bind("<Return>", lambda e: self.search_comments())
        
        search_btn = tk.Button(
            search_frame, 
            text="Buscar", 
            bg=BURDEOS, 
            fg="white",
            command=self.search_comments
        )
        search_btn.pack(side=tk.LEFT, padx=5)
        
        # Lista de comentarios
        self.comments_listbox = tk.Listbox(
            main_frame,
            height=15,
            selectmode=tk.SINGLE,
            font=("Arial", 10)
        )
        self.comments_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Scrollbar para la lista
        scrollbar = tk.Scrollbar(self.comments_listbox)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.comments_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.comments_listbox.yview)
        
        # Frame para botones de acción
        action_frame = tk.Frame(main_frame)
        action_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(0, 5))
        
        # Botón para ver comentario
        view_btn = tk.Button(
            action_frame, 
            text="Editar Comentario", 
            bg=BURDEOS, 
            fg="white",
            command=self.view_comment
        )
        view_btn.pack(side=tk.LEFT, padx=5)
        
        # Botón para eliminar comentario
        delete_btn = tk.Button(
            action_frame, 
            text="Eliminar Comentario", 
            bg="#8B0000", 
            fg="white",
            command=self.delete_comment
        )
        delete_btn.pack(side=tk.LEFT, padx=5)
        
        # Botón para cerrar
        close_btn = tk.Button(
            action_frame, 
            text="Cerrar", 
            bg="#606060", 
            fg="white",
            command=self.destroy
        )
        close_btn.pack(side=tk.RIGHT)
        
        # Cargar todos los comentarios inicialmente
        self.load_all_comments()
    
    def load_all_comments(self):
        """Carga todos los comentarios en la lista"""
        self.comments_listbox.delete(0, tk.END)
        if hasattr(self.editor, 'comments_data') and self.editor.comments_data:
            for comment_id, content in self.editor.comments_data.items():
                short_content = (content[:50] + '...') if len(content) > 50 else content
                self.comments_listbox.insert(tk.END, f"{comment_id}: {short_content}")
    
    def search_comments(self, event=None):
        """Busca comentarios que coincidan con el texto"""
        search_term = self.search_entry.get().strip().lower()
        self.comments_listbox.delete(0, tk.END)
        
        if not hasattr(self.editor, 'comments_data') or not self.editor.comments_data:
            return
        
        for comment_id, content in self.editor.comments_data.items():
            if not search_term or search_term in content.lower():
                short_content = (content[:50] + '...') if len(content) > 50 else content
                self.comments_listbox.insert(tk.END, f"{comment_id}: {short_content}")
    
    def get_selected_comment_id(self):
        """Obtiene el ID del comentario seleccionado"""
        selection = self.comments_listbox.curselection()
        if not selection:
            return None
        
        selected_text = self.comments_listbox.get(selection[0])
        # Extraer el ID del comentario del texto mostrado
        comment_id = selected_text.split(":")[0].strip()
        return comment_id if comment_id in self.editor.comments_data else None
    
    def view_comment(self):
        """Abre el comentario seleccionado para edición"""
        comment_id = self.get_selected_comment_id()
        if comment_id:
            content = self.editor.comments_data[comment_id]
            EditCommentWindow(
                self,
                content,
                callback=lambda updated_content: self.update_comment(comment_id, updated_content)
            )
    
    def update_comment(self, comment_id, updated_content):
        """Actualiza el comentario en el editor"""
        if comment_id in self.editor.comments_data:
            self.editor.comments_data[comment_id] = updated_content
            # Actualizar la lista
            self.search_comments()
    
    def delete_comment(self):
        """Elimina el comentario seleccionado"""
        comment_id = self.get_selected_comment_id()
        if not comment_id:
            messagebox.showwarning("Eliminar Comentario", "Por favor seleccione un comentario para eliminar")
            return
        
        if messagebox.askyesno(
            "Confirmar Eliminación", 
            f"¿Está seguro que desea eliminar el comentario {comment_id}?\nEsta acción no se puede deshacer."
        ):
            # Eliminar el marcador del texto
            marker = f"[Comentario{comment_id.split('_')[1]}]"
            start_pos = "1.0"
            
            while True:
                start_pos = self.editor.text_widget.search(marker, start_pos, stopindex=tk.END)
                if not start_pos:
                    break
                end_pos = f"{start_pos}+{len(marker)}c"
                self.editor.text_widget.delete(start_pos, end_pos)
                start_pos = end_pos
            
            # Eliminar el comentario de los datos
            del self.editor.comments_data[comment_id]
            
            # Actualizar la lista
            self.search_comments()
            
            messagebox.showinfo("Comentario Eliminado", "El comentario ha sido eliminado exitosamente")


class TomoEditor(tk.Frame):
    def __init__(self, master, tomo_name, **kwargs):
        super().__init__(master, **kwargs)
        self.tomo_name = tomo_name
        self.configure(bg='white')
        
        self.page_width = 670
        self.page_height = 940
        self.margin_top = 125
        self.margin_bottom = 64
        self.margin_left = 54
        self.margin_right = 15

        self.writeable_height = self.page_height - self.margin_top - self.margin_bottom
        self.writeable_width = self.page_width - self.margin_left - self.margin_right

        self.page_frame = tk.Frame(self, bg="#f9f9f9", relief="ridge", bd=3,
                                   width=self.page_width, height=self.page_height)
        self.page_frame.pack(expand=True, pady=10)

        self.bg_image = None
        self.load_bg_image()

        self.create_page()

    def load_bg_image(self):
        try:
            if os.path.exists("Hoja.png"):
                bg_image_orig = Image.open("Hoja.png").resize((self.page_width, self.page_height), Image.LANCZOS)
                self.bg_image = ImageTk.PhotoImage(bg_image_orig)
            else:
                bg_image_orig = Image.new("RGB", (self.page_width, self.page_height), "white")
                self.bg_image = ImageTk.PhotoImage(bg_image_orig)
        except Exception as e:
            print(f"Error cargando imagen de fondo: {e}")
            bg_image_orig = Image.new("RGB", (self.page_width, self.page_height), "white")
            self.bg_image = ImageTk.PhotoImage(bg_image_orig)

    def create_page(self):
        if not self.bg_image:
            self.load_bg_image()
            
        bg_label = tk.Label(self.page_frame, image=self.bg_image)
        bg_label.place(x=0, y=0, width=self.page_width, height=self.page_height)

        text_frame = tk.Frame(self.page_frame)
        text_frame.place(x=self.margin_left, y=self.margin_top, 
                         width=self.writeable_width, height=self.writeable_height)

        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.base_font = tkfont.Font(family="Arial", size=11)

        self.text_widget = tk.Text(
            text_frame,
            wrap="word",
            undo=True,
            font=self.base_font,
            relief="flat",
            bg="#ffffff",
            borderwidth=0,
            padx=0,
            pady=0,
            yscrollcommand=scrollbar.set
        )

        scrollbar.config(command=self.text_widget.yview)
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configuración visual
        self.text_widget.tag_configure("search", background="yellow")
        self.text_widget.tag_configure("bold", font=(self.base_font.actual("family"), self.base_font.actual("size"), "bold"))
        self.text_widget.tag_configure("italic", font=(self.base_font.actual("family"), self.base_font.actual("size"), "italic"))
        self.text_widget.tag_configure("underline", font=(self.base_font.actual("family"), self.base_font.actual("size"), "underline"))
        self.text_widget.tag_configure("comment", font=("Arial", 9, "italic"), foreground="blue")

        self.comments_data = {}

    def insert_comment(self, content):
        """Inserta un nuevo comentario"""
        if content:
            if not hasattr(self, "comment_counter"):
                self.comment_counter = 1
            else:
                self.comment_counter += 1

            comment_id = f"comment_{self.comment_counter}"
            marker = f"[Comentario{self.comment_counter}]"
            self.text_widget.insert(tk.INSERT, marker, comment_id)

            self.text_widget.tag_configure(comment_id, foreground="blue", font=("Arial", 9, "italic underline"))
            self.text_widget.tag_bind(comment_id, "<Enter>", lambda e: self.text_widget.config(cursor="hand2"))
            self.text_widget.tag_bind(comment_id, "<Leave>", lambda e: self.text_widget.config(cursor=""))
            self.text_widget.tag_bind(comment_id, "<Button-1>", lambda e, cid=comment_id: self.show_comment(cid))

            self.comments_data[comment_id] = content

    def show_comment(self, comment_id):
        """Permite ver y editar el comentario asociado al tag"""
        if hasattr(self, "comments_data") and comment_id in self.comments_data:
            EditCommentWindow(
                self,
                self.comments_data[comment_id],
                callback=lambda updated: self.comments_data.update({comment_id: updated})
            )

    def get_full_text(self):
        """Obtiene el texto tal como está escrito, más los comentarios si existen."""
        final_text = self.text_widget.get("1.0", tk.END).rstrip()

        if self.comments_data:
            final_text += "\n[[comentarios]]\n"
            for k, v in self.comments_data.items():
                final_text += f"{k}={v}\n"
            final_text += "[[/comentarios]]"

        return final_text

    def set_full_text(self, content):
        """Carga el texto completo, restaura los comentarios y el formato de temas."""
        self.text_widget.delete("1.0", tk.END)
        self.comments_data = {}

        # Extraer bloque de comentarios
        match = re.search(r"\[\[comentarios\]\](.*?)\[\[/comentarios\]\]", content, re.DOTALL)
        if match:
            comments_block = match.group(1).strip()
            for line in comments_block.splitlines():
                if "=" in line:
                    k, v = line.split("=", 1)
                    self.comments_data[k.strip()] = v.strip()

            # Actualizar contador después de cargar los comentarios
            existing_numbers = [
                int(k.split("_")[1]) for k in self.comments_data.keys() 
                if k.startswith("comment_") and k.split("_")[1].isdigit()
            ]
            self.comment_counter = max(existing_numbers) if existing_numbers else 0
            content = re.sub(r"\[\[comentarios\]\].*?\[\[/comentarios\]\]", "", content, flags=re.DOTALL)

        # Restaurar formato de temas
        # Insertar el contenido de texto plano (sin parsear temas especiales)
        self.text_widget.insert(tk.END, content.strip())

        # Restaurar cada marcador de comentario y sus eventos
        for comment_id in self.comments_data:
            comment_number = comment_id.replace("comment_", "")
            marker = f"[Comentario{comment_number}]"
            start = "1.0"
            while True:
                start = self.text_widget.search(marker, start, stopindex=tk.END)
                if not start:
                    break
                end = f"{start}+{len(marker)}c"
                self.text_widget.tag_add(comment_id, start, end)
                self.text_widget.tag_configure(comment_id, foreground="blue", font=("Arial", 9, "italic underline"))
                self.text_widget.tag_bind(comment_id, "<Enter>", lambda e: self.text_widget.config(cursor="hand2"))
                self.text_widget.tag_bind(comment_id, "<Leave>", lambda e: self.text_widget.config(cursor=""))
                self.text_widget.tag_bind(comment_id, "<Button-1>", lambda e, cid=comment_id: self.show_comment(cid))
                start = end

    def search_text(self, search_term, direction="next"):
        """Busca texto en el editor y resalta coincidencias"""
        # Limpiar resaltados anteriores
        self.text_widget.tag_remove("search", "1.0", tk.END)
        
        if not search_term:
            return
        
        # Empezar desde el cursor actual
        start_index = "1.0"
        if direction == "next":
            start_index = self.text_widget.index(tk.INSERT)
        elif direction == "prev":
            start_index = self.text_widget.index(tk.INSERT + " -1c")
        
        # Buscar coincidencias
        matches = []
        index = "1.0"
        while True:
            index = self.text_widget.search(search_term, index, stopindex=tk.END, 
                                          nocase=True, regexp=False)
            if not index:
                break
            end_index = f"{index}+{len(search_term)}c"
            
            # Resaltar coincidencia
            self.text_widget.tag_add("search", index, end_index)
            
            # Guardar posición para navegación
            matches.append((index, end_index))
            
            index = end_index
        
        if not matches:
            return None
        
        # Navegar a la siguiente/anterior coincidencia
        if direction == "next":
            # Encontrar la primera coincidencia después del cursor
            for match in matches:
                if self.text_widget.compare(match[0], ">=", tk.INSERT):
                    self.text_widget.see(match[0])
                    self.text_widget.mark_set(tk.INSERT, match[0])
                    return match
            # Si no hay coincidencias posteriores, ir a la primera
            self.text_widget.see(matches[0][0])
            self.text_widget.mark_set(tk.INSERT, matches[0][0])
            return matches[0]
        
        elif direction == "prev":
            # Encontrar la última coincidencia antes del cursor
            for i in range(len(matches)-1, -1, -1):
                if self.text_widget.compare(matches[i][0], "<=", tk.INSERT):
                    self.text_widget.see(matches[i][0])
                    self.text_widget.mark_set(tk.INSERT, matches[i][0])
                    return matches[i]
            # Si no hay coincidencias anteriores, ir a la última
            last_match = matches[-1]
            self.text_widget.see(last_match[0])
            self.text_widget.mark_set(tk.INSERT, last_match[0])
            return last_match

    def go_to_end(self):
        """Mueve el cursor al final del texto y desplaza la vista"""
        self.text_widget.mark_set(tk.INSERT, tk.END)
        self.text_widget.see(tk.END)
        
    def insert_theme(self, theme):
        """Inserta un tema como texto plano seguido de salto de línea."""
        index = self.text_widget.index(tk.INSERT)
        self.text_widget.insert(index, theme + "\n")
        self.text_widget.mark_set(tk.INSERT, f"{index}+{len(theme) + 1}c")
        self.text_widget.focus_set()
        self.text_widget.see(tk.INSERT)
        
    def apply_format(self, tag_name):
        """Aplica o remueve formato al texto seleccionado"""
        try:
            start_index = self.text_widget.index(tk.SEL_FIRST)
            end_index = self.text_widget.index(tk.SEL_LAST)
            
            # Verificar si el tag ya está presente en el rango
            tag_ranges = self.text_widget.tag_ranges(tag_name)
            tag_present = False
            for i in range(0, len(tag_ranges), 2):
                if (self.text_widget.compare(tag_ranges[i], "<=", start_index) and
                    self.text_widget.compare(tag_ranges[i+1], ">=", end_index)):
                    tag_present = True
                    break
            
            if tag_present:
                # Quitar el tag si está presente
                self.text_widget.tag_remove(tag_name, start_index, end_index)
            else:
                # Aplicar el tag si no está presente
                self.text_widget.tag_add(tag_name, start_index, end_index)
                
        except tk.TclError:
            # No hay texto seleccionado
            pass


class DatabaseManager:
    """Clase para manejar operaciones de base de datos"""
    def __init__(self, db_name="cuadernos.db"):
        self.db_name = db_name
        self.conn = None
        self.create_connection()
        self.create_tables()

    def create_connection(self):
        """Crear conexión a la base de datos"""
        try:
            self.conn = sqlite3.connect(self.db_name)
        except sqlite3.Error as e:
            print(f"Error conectando a la base de datos: {e}")

    def create_tables(self):
        """Crear tablas si no existen"""
        if self.conn is not None:
            try:
                cursor = self.conn.cursor()
                
                # Verificar si las tablas ya existen
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='cuadernos'")
                if not cursor.fetchone():
                    # Crear tabla cuadernos si no existe
                    cursor.execute('''
                        CREATE TABLE cuadernos (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            nombre TEXT NOT NULL UNIQUE,
                            fecha_creacion TEXT,
                            fecha_modificacion TEXT
                        )
                    ''')
                
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='hojas'")
                if not cursor.fetchone():
                    # Crear tabla hojas si no existe
                    cursor.execute('''
                        CREATE TABLE hojas (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            cuaderno_id INTEGER NOT NULL,
                            contenido TEXT,
                            fecha_creacion TEXT,
                            fecha_modificacion TEXT,
                            FOREIGN KEY (cuaderno_id) REFERENCES cuadernos (id)
                        )
                    ''')
                
                # Verificar si la columna cuaderno_id existe en hojas
                cursor.execute("PRAGMA table_info(hojas)")
                columns = [column[1] for column in cursor.fetchall()]
                if "cuaderno_id" not in columns:
                    # Si la columna no existe, recrear la tabla
                    cursor.execute("DROP TABLE IF EXISTS hojas")
                    cursor.execute('''
                        CREATE TABLE hojas (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            cuaderno_id INTEGER NOT NULL,
                            contenido TEXT,
                            fecha_creacion TEXT,
                            fecha_modificacion TEXT,
                            FOREIGN KEY (cuaderno_id) REFERENCES cuadernos (id)
                        )
                    ''')
                
                self.conn.commit()
            except sqlite3.Error as e:
                print(f"Error creando tablas: {e}")

    def get_cuaderno_id(self, nombre):
        """Obtener ID de un cuaderno por nombre, crearlo si no existe"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT id FROM cuadernos WHERE nombre = ?", (nombre,))
        result = cursor.fetchone()
        
        if result:
            return result[0]
        else:
            # Crear nuevo cuaderno
            now = datetime.now().isoformat()
            cursor.execute(
                "INSERT INTO cuadernos (nombre, fecha_creacion, fecha_modificacion) VALUES (?, ?, ?)",
                (nombre, now, now)
            )
            self.conn.commit()
            return cursor.lastrowid

    def guardar_hoja(self, cuaderno_nombre, contenido):
        """Guardar o actualizar una hoja en la base de datos"""
        cuaderno_id = self.get_cuaderno_id(cuaderno_nombre)
        cursor = self.conn.cursor()
        now = datetime.now().isoformat()
        
        # Verificar si ya existe una hoja para este cuaderno
        cursor.execute("SELECT id FROM hojas WHERE cuaderno_id = ?", (cuaderno_id,))
        existing = cursor.fetchone()
        
        if existing:
            # Actualizar hoja existente
            cursor.execute(
                "UPDATE hojas SET contenido = ?, fecha_modificacion = ? WHERE id = ?",
                (contenido, now, existing[0])
            )
        else:
            # Crear nueva hoja
            cursor.execute(
                "INSERT INTO hojas (cuaderno_id, contenido, fecha_creacion, fecha_modificacion) VALUES (?, ?, ?, ?)",
                (cuaderno_id, contenido, now, now)
            )
        
        self.conn.commit()

    def cargar_hoja(self, cuaderno_nombre):
        """Cargar contenido de una hoja desde la base de datos"""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT contenido FROM hojas WHERE cuaderno_id = (SELECT id FROM cuadernos WHERE nombre = ?)",
            (cuaderno_nombre,)
        )
        result = cursor.fetchone()
        return result[0] if result else ""

    def cerrar_conexion(self):
        """Cerrar conexión a la base de datos"""
        if self.conn:
            self.conn.close()


class TextEditorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Editor de Cuadernos Regionales")
        self.root.geometry("1200x800")  # Mayor ancho para la barra de tomos

        # Inicializar base de datos
        self.db_manager = DatabaseManager()

        # Lista de tomos con sus nombres
        self.tomo_names = [
            "Arica", "Tarapacá", "Antofagasta", "Atacama", "Coquimbo", 
            "Valparaíso", "R. Metropolitana", "O'Higgins", "Maule", "Ñuble", 
            "Bío-Bío", "Araucanía", "Los Ríos", "Los Lagos", "Aysén", 
            "Magallanes", "General"
        ]
        
        # Lista de temas predefinidos
        self.temas = [
            "Clima Laboral",
            "Ejecución Presupuestaria",
            "Indicadores de desempeño",
            "Informática",
            "Infraestructura",
            "Planificación",
            "Plan de SSPP",
            "Político Institucional",
            "Otros",
            "Temas Dpto. Personas",
        ]
        
        self.tomos = {}  # Diccionario para guardar los editores de cada tomo
        self.current_tomo = None
        self.last_theme_position = {}  # Para rastrear la posición del último tema por tomo
        self.last_theme_tema = {}     # Para rastrear el último tema insertado por tomo
        
        # Crear paneles principales
        main_paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=5)
        main_paned.pack(fill=tk.BOTH, expand=True)
        
        # Panel izquierdo: barra de tomos
        tomo_frame = tk.Frame(main_paned, width=200, bg=BURDEOS)
        main_paned.add(tomo_frame)
        
        # Título de la barra de tomos
        title_label = tk.Label(
            tomo_frame, 
            text="  CUADERNOS DE NOTAS  ", 
            font=("Arial", 10, "bold"), 
            bg=BURDEOS, 
            fg="white",
            pady=10
        )
        title_label.pack(fill=tk.X)
        
        # Crear botones para cada tomo
        self.tomo_buttons = []
        for name in self.tomo_names:
            btn = tk.Button(
                tomo_frame, 
                text=name, 
                width=15, 
                height=2, 
                bg=BURDEOS, 
                fg="white", 
                font=("Arial", 10, "bold"),
                relief="flat", 
                activebackground="#a04060",
                command=lambda n=name: self.show_tomo(n)
            )
            btn.pack(pady=5, padx=10, fill=tk.X)
            self.tomo_buttons.append(btn)
            # Inicializar rastreo de posición para cada tomo
            self.last_theme_position[name] = None
            self.last_theme_tema[name] = None
        
        # Panel derecho: área de edición y búsqueda
        editor_search_frame = tk.Frame(main_paned)
        main_paned.add(editor_search_frame, width=1000)
        
        # Barra de herramientas principal
        toolbar = tk.Frame(editor_search_frame, bg="#f0f0f0", padx=5, pady=5)
        toolbar.pack(fill=tk.X)

        # Grupo 1: Búsqueda y navegación
        search_group = tk.Frame(toolbar, bg="#f0f0f0")
        search_group.pack(side=tk.LEFT, padx=(0, 10))

        tk.Label(search_group, text="Buscar:", bg="#f0f0f0").pack(side=tk.LEFT)

        self.search_entry = tk.Entry(search_group, width=25)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind("<Return>", lambda e: self.search_text("next"))

        self.search_prev_btn = tk.Button(
            search_group, 
            text="◄", 
            bg=BURDEOS, 
            fg="white",
            width=3,
            command=lambda: self.search_text("prev")
        )
        self.search_prev_btn.pack(side=tk.LEFT, padx=(5, 0))

        self.search_next_btn = tk.Button(
            search_group, 
            text="►", 
            bg=BURDEOS, 
            fg="white",
            width=3,
            command=lambda: self.search_text("next")
        )
        self.search_next_btn.pack(side=tk.LEFT)

        self.clear_search_btn = tk.Button(
            search_group, 
            text="✕", 
            bg=GRIS_OSCURO, 
            fg="white",
            width=3,
            command=self.clear_search
        )
        self.clear_search_btn.pack(side=tk.LEFT, padx=(5, 0))

        # Botón Fin
        self.goto_end_btn = tk.Button(
            search_group, 
            text="Fin", 
            bg=GRIS_OSCURO, 
            fg="white",
            width=5,
            command=self.go_to_end
        )
        self.goto_end_btn.pack(side=tk.LEFT, padx=(10, 0))

        # Separador visual
        tk.Frame(toolbar, bg="#d0d0d0", width=1, height=28).pack(side=tk.LEFT, padx=10)

        # Grupo 2: Temas y comentarios
        theme_group = tk.Frame(toolbar, bg="#f0f0f0")
        theme_group.pack(side=tk.LEFT, padx=(0, 10))

        # Combobox para temas
        self.theme_combo = ttk.Combobox(
            theme_group, 
            width=22,
            values=self.temas,
            state="readonly"
        )
        self.theme_combo.pack(side=tk.LEFT)
        self.theme_combo.bind("<<ComboboxSelected>>", self.insert_selected_theme)

        self.insert_theme_btn = tk.Button(
            theme_group, 
            text="Tema", 
            bg=AZUL_OSCURO, 
            fg="white",
            width=6,
            command=self.insert_selected_theme
        )
        self.insert_theme_btn.pack(side=tk.LEFT, padx=5)

        # Separador visual dentro del grupo
        tk.Frame(theme_group, bg="#d0d0d0", width=1, height=28).pack(side=tk.LEFT, padx=5)

        # Grupo 3: Comentarios
        comment_group = tk.Frame(toolbar, bg="#f0f0f0")
        comment_group.pack(side=tk.LEFT, padx=(0, 10))

        self.insert_comment_btn = tk.Button(
            comment_group, 
            text="Comentario", 
            bg=AZUL_OSCURO, 
            fg="white",
            width=10,
            command=self.insert_comment
        )
        self.insert_comment_btn.pack(side=tk.LEFT)

        self.search_comments_btn = tk.Button(
            comment_group, 
            text="Buscar Comentarios", 
            bg=BURDEOS, 
            fg="white",
            width=16,
            command=self.search_comments
        )
        self.search_comments_btn.pack(side=tk.LEFT, padx=5)

        # Separador visual
        tk.Frame(toolbar, bg="#d0d0d0", width=1, height=28).pack(side=tk.LEFT, padx=10)

        # Grupo 4: Formato
        format_group = tk.Frame(toolbar, bg="#f0f0f0")
        format_group.pack(side=tk.LEFT)

        self.bold_btn = tk.Button(
            format_group, 
            text="N", 
            bg=BURDEOS, 
            fg="white",
            width=3,
            command=lambda: self.apply_format("bold")
        )
        self.bold_btn.pack(side=tk.LEFT)

        self.italic_btn = tk.Button(
            format_group, 
            text="K", 
            bg=BURDEOS, 
            fg="white",
            width=3,
            command=lambda: self.apply_format("italic")
        )
        self.italic_btn.pack(side=tk.LEFT, padx=5)

        self.underline_btn = tk.Button(
            format_group, 
            text="S", 
            bg=BURDEOS, 
            fg="white",
            width=3,
            command=lambda: self.apply_format("underline")
        )
        self.underline_btn.pack(side=tk.LEFT)
        
        # Área de edición
        self.editor_frame = tk.Frame(editor_search_frame)
        self.editor_frame.pack(fill=tk.BOTH, expand=True)
        
        # Mostrar el primer tomo por defecto
        self.show_tomo(self.tomo_names[0])

        self.file_path = None

        self.create_menu()
        self.bind_shortcuts()

    def show_tomo(self, tomo_name):
        """Muestra el editor del tomo seleccionado"""
        # Crear el editor si no existe
        if tomo_name not in self.tomos:
            self.tomos[tomo_name] = TomoEditor(self.editor_frame, tomo_name)
            # Cargar contenido desde base de datos
            contenido = self.db_manager.cargar_hoja(tomo_name)
            self.tomos[tomo_name].set_full_text(contenido)
            self.tomos[tomo_name].pack(fill=tk.BOTH, expand=True)
        
        # Ocultar todos los tomos y mostrar el actual
        for name, editor in self.tomos.items():
            if name == tomo_name:
                editor.pack(fill=tk.BOTH, expand=True)
                self.current_tomo = tomo_name
                self.root.title(f"Editor de Cuadernos Regionales - {tomo_name}")
            else:
                editor.pack_forget()
        
        # Actualizar estado de los botones
        for btn in self.tomo_buttons:
            if btn["text"] == tomo_name:
                btn.config(relief="sunken", bg="#a04060")
            else:
                btn.config(relief="flat", bg=BURDEOS)
                
        # Limpiar búsqueda al cambiar de tomo
        self.clear_search()

    def search_text(self, direction):
        """Busca texto en el editor actual"""
        search_term = self.search_entry.get().strip()
        if not search_term:
            return
            
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            match = editor.search_text(search_term, direction)
            
            if not match:
                messagebox.showinfo("Búsqueda", "No se encontraron coincidencias")

    def clear_search(self):
        """Limpia los resultados de búsqueda"""
        self.search_entry.delete(0, tk.END)
        if self.current_tomo and self.current_tomo in self.tomos:
            self.tomos[self.current_tomo].text_widget.tag_remove("search", "1.0", tk.END)

    def go_to_end(self):
        """Mueve el cursor al final del texto en el editor actual"""
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            editor.go_to_end()
            
    def insert_comment(self):
        """Inserta un nuevo comentario en el editor actual"""
        if self.current_tomo and self.current_tomo in self.tomos:
            # Abrir ventana para crear comentario
            CommentWindow(self.root, callback=lambda content: self.save_comment(content))
            
    def save_comment(self, content):
        """Guarda el comentario en el editor actual"""
        if self.current_tomo and self.current_tomo in self.tomos and content:
            editor = self.tomos[self.current_tomo]
            editor.insert_comment(content)
            
    def search_comments(self):
        """Abre la ventana de búsqueda y gestión de comentarios"""
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            CommentSearchWindow(self.root, editor)
            
    def insert_selected_theme(self, event=None):
        """Inserta el tema seleccionado en el editor activo"""
        selected_theme = self.theme_combo.get()
        if not selected_theme:
            return
            
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            
            # Guardar posición actual del cursor
            current_position = editor.text_widget.index(tk.INSERT)
            
            # Insertar el tema
            editor.insert_theme(selected_theme)
            
            # Si hay un tema anterior, guardar el detalle en Excel
            if self.last_theme_position[self.current_tomo]:
                start_index = self.last_theme_position[self.current_tomo]
                end_index = current_position
                detalle = editor.text_widget.get(start_index, end_index).strip()
                
                # Guardar en Excel solo si hay detalle
                if detalle:
                    ExcelManager.save_to_excel(
                        tomo_name=self.current_tomo,
                        tema=self.last_theme_tema[self.current_tomo],
                        detalle=detalle
                    )
            
            # Actualizar la posición para el próximo detalle
            self.last_theme_position[self.current_tomo] = editor.text_widget.index(tk.INSERT)
            self.last_theme_tema[self.current_tomo] = selected_theme
            
            # Limpiar la selección después de insertar
            self.theme_combo.set('')

    def apply_format(self, format_type):
        """Aplica formato al texto seleccionado"""
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            editor.apply_format(format_type)

    def create_menu(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Nuevo", command=self.new_file)
        file_menu.add_command(label="Abrir...", command=self.open_file)
        file_menu.add_command(label="Guardar", command=self.save_file)
        file_menu.add_command(label="Guardar como...", command=self.save_file_as)
        file_menu.add_separator()
        file_menu.add_command(label="Salir", command=self.on_exit)
        menu_bar.add_cascade(label="Archivo", menu=file_menu)

        edit_menu = tk.Menu(menu_bar, tearoff=0)
        edit_menu.add_command(label="Deshacer", command=self.undo)
        edit_menu.add_command(label="Rehacer", command=self.redo)
        edit_menu.add_separator()
        edit_menu.add_command(label="Cortar", command=self.cut)
        edit_menu.add_command(label="Copiar", command=self.copy)
        edit_menu.add_command(label="Pegar", command=self.paste)
        edit_menu.add_separator()
        edit_menu.add_command(label="Ir al final", command=self.go_to_end)
        edit_menu.add_command(label="Seleccionar todo", command=self.select_all)
        edit_menu.add_separator()
        edit_menu.add_command(label="Buscar...", command=self.focus_search)
        edit_menu.add_command(label="Buscar Comentarios...", command=self.search_comments)
        menu_bar.add_cascade(label="Edición", menu=edit_menu)
        
        # Menú para temas
        theme_menu = tk.Menu(menu_bar, tearoff=0)
        for tema in self.temas:
            theme_menu.add_command(
                label=tema, 
                command=lambda t=tema: self.insert_theme_from_menu(t)
            )
        menu_bar.add_cascade(label="Temas", menu=theme_menu)
        
        # Menú para formato
        format_menu = tk.Menu(menu_bar, tearoff=0)
        format_menu.add_command(label="Negrita", command=lambda: self.apply_format("bold"), accelerator="Ctrl+B")
        format_menu.add_command(label="Cursiva", command=lambda: self.apply_format("italic"), accelerator="Ctrl+I")
        format_menu.add_command(label="Subrayado", command=lambda: self.apply_format("underline"), accelerator="Ctrl+U")
        format_menu.add_command(label="Comentario", command=self.insert_comment, accelerator="Ctrl+M")
        menu_bar.add_cascade(label="Formato", menu=format_menu)

    def insert_theme_from_menu(self, tema):
        """Inserta un tema desde el menú"""
        if self.current_tomo and self.current_tomo in self.tomos:
            # Usar el mismo método que para el combobox
            self.theme_combo.set(tema)
            self.insert_selected_theme()

    def on_exit(self):
        """Manejar salida de la aplicación"""
        if hasattr(self, 'db_manager'):
            self.db_manager.cerrar_conexion()
        self.root.quit()

    def focus_search(self):
        """Enfoca el campo de búsqueda"""
        self.search_entry.focus_set()

    def bind_shortcuts(self):
        self.root.bind_all("<Control-n>", lambda e: self.new_file())
        self.root.bind_all("<Control-o>", lambda e: self.open_file())
        self.root.bind_all("<Control-s>", lambda e: self.save_file())
        self.root.bind_all("<Control-Shift-S>", lambda e: self.save_file_as())
        self.root.bind_all("<Control-a>", lambda e: self.select_all())
        self.root.bind_all("<Control-z>", lambda e: self.undo())
        self.root.bind_all("<Control-y>", lambda e: self.redo())
        self.root.bind_all("<Control-x>", lambda e: self.cut())
        self.root.bind_all("<Control-c>", lambda e: self.copy())
        self.root.bind_all("<Control-v>", lambda e: self.paste())
        self.root.bind_all("<Control-f>", lambda e: self.focus_search())
        self.root.bind_all("<Control-e>", lambda e: self.go_to_end())
        self.root.bind_all("<Control-m>", lambda e: self.insert_comment())
        # Atajos para formato
        self.root.bind_all("<Control-b>", lambda e: self.apply_format("bold"))
        self.root.bind_all("<Control-i>", lambda e: self.apply_format("italic"))
        self.root.bind_all("<Control-u>", lambda e: self.apply_format("underline"))
        # Nuevo atajo para buscar comentarios
        self.root.bind_all("<Control-Shift-M>", lambda e: self.search_comments())

    def get_current_widget(self):
        """Obtiene el widget de texto actual del tomo actual"""
        if self.current_tomo and self.current_tomo in self.tomos:
            return self.tomos[self.current_tomo].text_widget
        return None

    def undo(self):
        widget = self.get_current_widget()
        if widget:
            try:
                widget.edit_undo()
            except tk.TclError:
                pass

    def redo(self):
        widget = self.get_current_widget()
        if widget:
            try:
                widget.edit_redo()
            except tk.TclError:
                pass

    def cut(self):
        widget = self.get_current_widget()
        if widget:
            widget.event_generate("<<Cut>>")

    def copy(self):
        widget = self.get_current_widget()
        if widget:
            widget.event_generate("<<Copy>>")

    def paste(self):
        widget = self.get_current_widget()
        if widget:
            widget.event_generate("<<Paste>>")

    def select_all(self):
        widget = self.get_current_widget()
        if widget:
            widget.tag_add(tk.SEL, "1.0", tk.END)
            widget.mark_set(tk.INSERT, "1.0")
            widget.see(tk.INSERT)

    def new_file(self):
        if self.confirm_unsaved_changes():
            # Limpiar el tomo actual
            if self.current_tomo and self.current_tomo in self.tomos:
                self.tomos[self.current_tomo].set_full_text("")
            self.file_path = None
            self.root.title(f"Editor de Cuadernos Regionales - {self.current_tomo}")

    def open_file(self):
        if not self.confirm_unsaved_changes():
            return
        file_path = filedialog.askopenfilename(filetypes=[("Archivos de texto", "*.txt")])
        if file_path:
            with open(file_path, "r", encoding="utf-8") as file:
                content = file.read()
            # Cargar en el tomo actual
            if self.current_tomo and self.current_tomo in self.tomos:
                self.tomos[self.current_tomo].set_full_text(content)
            self.file_path = file_path
            self.root.title(f"Editor de Cuadernos Regionales - {self.current_tomo} - {file_path}")

    def save_file(self):
        """Guarda el contenido actual en la base de datos"""
        if self.current_tomo and self.current_tomo in self.tomos:
            editor = self.tomos[self.current_tomo]
            content = editor.get_full_text()
            
            # Guardar en base de datos
            self.db_manager.guardar_hoja(self.current_tomo, content)
            
            # Guardar el último detalle si existe
            if self.last_theme_position.get(self.current_tomo):
                start_index = self.last_theme_position[self.current_tomo]
                detalle = editor.text_widget.get(start_index, tk.END).strip()
                
                if detalle and self.last_theme_tema.get(self.current_tomo):
                    ExcelManager.save_to_excel(
                        tomo_name=self.current_tomo,
                        tema=self.last_theme_tema[self.current_tomo],
                        detalle=detalle
                    )
            
            messagebox.showinfo("Guardado", f"Contenido de '{self.current_tomo}' guardado en la base de datos y Excel")
            # Marcar el widget como no modificado
            editor.text_widget.edit_modified(False)
        else:
            messagebox.showwarning("Guardar", "No hay un tomo seleccionado para guardar")

    def save_file_as(self):
        """Exporta el contenido actual a un archivo de texto"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Archivos de texto", "*.txt")],
            initialfile=f"{self.current_tomo}.txt" if self.current_tomo else "documento.txt"
        )
        if file_path:
            content = ""
            if self.current_tomo and self.current_tomo in self.tomos:
                content = self.tomos[self.current_tomo].get_full_text()
            with open(file_path, "w", encoding="utf-8") as file:
                file.write(content.rstrip())
            messagebox.showinfo("Exportado", f"Contenido exportado a {file_path}")

    def confirm_unsaved_changes(self):
        """Verifica si hay cambios sin guardar en el tomo actual"""
        if self.current_tomo and self.current_tomo in self.tomos:
            widget = self.tomos[self.current_tomo].text_widget
            if widget.edit_modified():
                response = messagebox.askyesnocancel(
                    "Guardar cambios", 
                    f"¿Deseas guardar los cambios en {self.current_tomo} antes de continuar?"
                )
                if response is None:  # Cancelar
                    return False
                elif response:  # Sí
                    self.save_file()
                else:  # No
                    widget.edit_modified(False)
        return True


if __name__ == "__main__":
    root = tk.Tk()
    app = TextEditorApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_exit)
    root.mainloop()