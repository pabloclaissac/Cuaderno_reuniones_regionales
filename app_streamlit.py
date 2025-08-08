import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np

# Configuración de la página
st.set_page_config(layout="wide", page_title="Editor de Cuadernos Regionales")

# Constantes y estilos
BURDEOS = "#ac042b"
AZUL_OSCURO = "#003366"
GRIS_OSCURO = "#333333"
EXCEL_FILE = "Planificación 2025.xlsx"
SHEET_NAME = "Hoja3"
DB_FILE = "cuadernos.db"

# Variables de ancho configurable
REGION_BUTTON_WIDTH = "200px"  # Ancho de botones de regiones
TOOLBAR_BUTTON_WIDTH = "20px"  # Ancho de botones en barra de herramientas
SEARCH_BUTTON_WIDTH = "20px"  # Ancho de botones de búsqueda
THEME_BUTTON_WIDTH = "20px"  # Ancho de botón "Insertar Tema"

# Nuevos estilos CSS para el diseño solicitado
st.markdown(f"""
    <style>
    /* Estilos generales de la barra lateral */
    [data-testid="stSidebar"] {{
        background-color: {BURDEOS};
        color: white;
    }}
    
    /* Título principal de la barra lateral */
    .sidebar-title {{
        font-size: 1.4rem;
        font-weight: bold;
        text-align: center;
        margin-bottom: 2rem;
        padding: 0.2rem;
        border-bottom: 2px solid rgba(255, 255, 255, 0.2);
    }}
    
    /* Contenedor de botones */
    .tomo-container {{
        display: flex;
        flex-direction: column;
        gap: 0.1rem;
        padding: 0 0.5rem;
        max-height: calc(100vh - 150px);
        overflow-y: auto;
    }}
    
    /* Botones de regiones - AHORA CON CLASE ESPECÍFICA */
    .region-button {{
        background-color: #ac042b;
        color: #ac042b !important;
        border: 1px solid rgba(255, 255, 255, 0.3);
        border-radius: 4px;
        padding: 0.7rem 0.5rem;
        text-align: center;
        font-size: 0.75rem;
        font-weight: bold;
        transition: all 0.3s ease;
        margin: 0;
        width: {REGION_BUTTON_WIDTH} !important;
        display: block !important;
        margin-left: auto !important;
        margin-right: auto !important;
    }}
    
    .region-button:hover {{
        background-color: #DDEFFB;
        transform: translateY(-1px);
        box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2);
    }}
    
    .active-tomo {{
        background-color: #c6254b !important;
        border: 1px solid rgba(255, 255, 255, 0.6);
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.3);
    }}
    
    /* Contenedor principal */
    .main-container {{
        display: flex;
        flex-direction: column;
        height: calc(5vh - 100px);
    }}
    
    /* Área superior (herramientas) */
    .toolbar-container {{
        background-color: white;
        padding: 3px;
        border-bottom: 1px solid #e0e0e0;
        margin-bottom: 5px;
    }}
    
    /* Área del editor */
    .editor-container {{
        flex-grow: 1;
        background-color: white;
        padding: 15px;
        overflow: hidden;
    }}
    
    /* Barra de herramientas */
    .toolbar-section {{
        margin-bottom: 0.2rem;
    }}
    
    /* Espaciado entre elementos */
    .spacer {{
        margin-top: 10px;
    }}
    
    /* Textarea del editor */
    .stTextArea textarea {{
        background-color: transparent !important;
        line-height: 1.5;
        font-size: 1.05rem;
        height: calc(100vh - 250px) !important;
    }}
    
    /* Encabezado principal */
    .main-header {{
        margin-bottom: 0.4rem;
        border-bottom: 2px solid {BURDEOS};
        padding-bottom: 0.4rem;
    }}
    
    /* Botones en barra de herramientas - CLASES ESPECÍFICAS */
    .toolbar-button {{
        width: {TOOLBAR_BUTTON_WIDTH} !important;
        box-sizing: border-box;
        display: block !important;
        margin-left: auto !important;
        margin-right: 0 !important;
    }}
    
    /* Botones de búsqueda - CLASES ESPECÍFICAS */
    .search-button-custom {{
        width: {SEARCH_BUTTON_WIDTH} !important;
        box-sizing: border-box;
    }}
    
    /* Botón Insertar Tema - CLASE ESPECÍFICA */
    .theme-button {{
        width: {THEME_BUTTON_WIDTH} !important;
        box-sizing: border-box;
        display: block !important;
        margin-left: auto !important;
        margin-right: auto !important;
    }}
    </style>
    """, unsafe_allow_html=True)

# Clase para manejar Excel
class ExcelManager:
    @staticmethod
    def get_next_empty_row(worksheet):
        row = 1
        while worksheet.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    @staticmethod
    def save_to_excel(tomo_name, tema, detalle):
        try:
            if not os.path.exists(EXCEL_FILE):
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                sheet = wb.create_sheet(SHEET_NAME)
                sheet['A1'] = "Dirección Regional"
                sheet['B1'] = "Fecha de Reunión"
                sheet['C1'] = "Ítem de monitoreo"
                sheet['D1'] = "Detalle"
            else:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                if SHEET_NAME in wb.sheetnames:
                    sheet = wb[SHEET_NAME]
                else:
                    sheet = wb.create_sheet(SHEET_NAME)
                    sheet['A1'] = "Dirección Regional"
                    sheet['B1'] = "Fecha de Reunión"
                    sheet['C1'] = "Ítem de monitoreo"
                    sheet['D1'] = "Detalle"
            
            next_row = ExcelManager.get_next_empty_row(sheet)
            sheet[f'A{next_row}'] = tomo_name
            sheet[f'B{next_row}'] = datetime.now().strftime("%d/%m/%Y")
            sheet[f'C{next_row}'] = tema
            sheet[f'D{next_row}'] = detalle
            wb.save(EXCEL_FILE)
            return True
        except Exception as e:
            st.error(f"Error al guardar en Excel: {e}")
            return False

# Clase para manejar la base de datos (conexión por hilo)
class DatabaseManager:
    def __init__(self, db_name=DB_FILE):
        self.db_name = db_name
        self.create_tables()
    
    def get_connection(self):
        """Crea una nueva conexión para cada operación"""
        return sqlite3.connect(self.db_name)
    
    def create_tables(self):
        """Crea las tablas si no existen"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cuadernos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE,
                fecha_creacion TEXT,
                fecha_modificacion TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS hojas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuaderno_id INTEGER NOT NULL,
                contenido TEXT,
                fecha_creacion TEXT,
                fecha_modificacion TEXT,
                FOREIGN KEY (cuaderno_id) REFERENCES cuadernos (id)
            )
        ''')
        conn.commit()
        conn.close()
    
    def get_cuaderno_id(self, nombre):
        """Obtiene el ID de un cuaderno, creándolo si no existe"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM cuadernos WHERE nombre = ?", (nombre,))
        result = cursor.fetchone()
        if result:
            conn.close()
            return result[0]
        else:
            now = datetime.now().isoformat()
            cursor.execute(
                "INSERT INTO cuadernos (nombre, fecha_creacion, fecha_modificacion) VALUES (?, ?, ?)",
                (nombre, now, now)
            )
            id = cursor.lastrowid
            conn.commit()
            conn.close()
            return id
    
    def guardar_hoja(self, cuaderno_nombre, contenido):
        """Guarda una hoja en la base de datos"""
        cuaderno_id = self.get_cuaderno_id(cuaderno_nombre)
        conn = self.get_connection()
        cursor = conn.cursor()
        now = datetime.now().isoformat()
        cursor.execute("SELECT id FROM hojas WHERE cuaderno_id = ?", (cuaderno_id,))
        existing = cursor.fetchone()
        if existing:
            cursor.execute(
                "UPDATE hojas SET contenido = ?, fecha_modificacion = ? WHERE id = ?",
                (contenido, now, existing[0])
            )
        else:
            cursor.execute(
                "INSERT INTO hojas (cuaderno_id, contenido, fecha_creacion, fecha_modificacion) VALUES (?, ?, ?, ?)",
                (cuaderno_id, contenido, now, now)
            )
        conn.commit()
        conn.close()
    
    def cargar_hoja(self, cuaderno_nombre):
        """Carga una hoja desde la base de datos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT contenido FROM hojas WHERE cuaderno_id = (SELECT id FROM cuadernos WHERE nombre = ?)",
            (cuaderno_nombre,))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else ""

# Inicialización de la aplicación
def init_session():
    if "db_manager" not in st.session_state:
        st.session_state.db_manager = DatabaseManager()
    
    if "tomo_names" not in st.session_state:
        st.session_state.tomo_names = [
            "Arica", "Tarapacá", "Antofagasta", "Atacama", "Coquimbo", 
            "Valparaíso", "R. Metropolitana", "O'Higgins", "Maule", "Ñuble", 
            "Bío-Bío", "Araucanía", "Los Ríos", "Los Lagos", "Aysén", 
            "Magallanes", "General"
        ]
    
    if "temas" not in st.session_state:
        st.session_state.temas = [
            "Clima Laboral", "Ejecución Presupuestaria", "Indicadores de desempeño",
            "Informática", "Infraestructura", "Planificación", "Plan de SSPP",
            "Político Institucional", "Otros", "Temas Dpto. Personas"
        ]
    
    if "current_tomo" not in st.session_state:
        st.session_state.current_tomo = st.session_state.tomo_names[0]
    
    if "contenido_tomo" not in st.session_state:
        st.session_state.contenido_tomo = st.session_state.db_manager.cargar_hoja(
            st.session_state.current_tomo)
    
    if "search_term" not in st.session_state:
        st.session_state.search_term = ""
    
    if "search_results" not in st.session_state:
        st.session_state.search_results = []
    
    if "current_search_index" not in st.session_state:
        st.session_state.current_search_index = 0

# Funciones de búsqueda
def search_text(direction):
    if not st.session_state.search_term:
        return
    
    content = st.session_state.contenido_tomo
    search_term = st.session_state.search_term.lower()
    matches = [m.start() for m in re.finditer(re.escape(search_term), content.lower())]
    
    if not matches:
        st.warning("No se encontraron coincidencias")
        return
    
    if direction == "next":
        if st.session_state.current_search_index < len(matches) - 1:
            st.session_state.current_search_index += 1
        else:
            st.session_state.current_search_index = 0
    else:  # prev
        if st.session_state.current_search_index > 0:
            st.session_state.current_search_index -= 1
        else:
            st.session_state.current_search_index = len(matches) - 1
    
    # Actualizar el editor para mostrar la coincidencia
    start_pos = matches[st.session_state.current_search_index]
    end_pos = start_pos + len(st.session_state.search_term)
    st.session_state.editor_area = content
    st.rerun()  # Cambiado de experimental_rerun() a rerun()

# Función para insertar temas
def insert_theme():
    theme = st.session_state.selected_theme
    if not theme:
        return
    
    # Insertar el tema en el contenido
    new_content = st.session_state.contenido_tomo + f"\n{theme}\n"
    st.session_state.contenido_tomo = new_content
    
    # Guardar en Excel
    ExcelManager.save_to_excel(
        st.session_state.current_tomo,
        theme,
        new_content)
    
    # Actualizar la base de datos
    st.session_state.db_manager.guardar_hoja(
        st.session_state.current_tomo, 
        new_content)
    
    st.rerun()

# Interfaz principal
def main():
    init_session()
    
    # Barra lateral (izquierda)
    with st.sidebar:
        st.markdown('<div class="sidebar-title">CUADERNOS DE NOTAS</div>', unsafe_allow_html=True)
        
        with st.container():
            st.markdown('<div class="tomo-container">', unsafe_allow_html=True)
            
            for tomo in st.session_state.tomo_names:
                # Botones de regiones con clase personalizada
                if st.button(
                    tomo, 
                    key=f"btn_{tomo}", 
                    use_container_width=True,
                    type="primary" if tomo == st.session_state.current_tomo else "secondary"
                ):
                    st.session_state.current_tomo = tomo
                    st.session_state.contenido_tomo = st.session_state.db_manager.cargar_hoja(tomo)
                    st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Área principal (derecha) con dos subáreas
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.markdown(f'<div class="main-header">Editor de Cuadernos Regionales - {st.session_state.current_tomo}</div>', 
               unsafe_allow_html=True)
    
    # Subárea superior (herramientas) - NUEVA ESTRUCTURA DE 3 COLUMNAS
    with st.container():
        st.markdown('<div class="toolbar-container">', unsafe_allow_html=True)
        
        # Dividir en 3 columnas de tamaño equivalente
        col1, col2, col3 = st.columns([1, 1, 1])
        
        # Columna 1: Búsqueda y controles
        with col1:
            st.session_state.search_term = st.text_input(
                "Buscar:", 
                st.session_state.search_term, 
                key="search_input",
                placeholder="Buscar en el documento..."
            )
            
            # Controles de búsqueda (en fila)
            st.markdown('<div class="spacer"></div>', unsafe_allow_html=True)
            search_col1, search_col2, search_col3 = st.columns([1, 1, 1])
            
            with search_col1:
                if st.button("◄ Anterior", key="btn_prev", use_container_width=True, 
                            help="Ir a la coincidencia anterior",
                            type="secondary"):
                    search_text("prev")
            
            with search_col2:
                if st.button("Siguiente ►", key="btn_next", use_container_width=True,
                            help="Ir a la siguiente coincidencia",
                            type="secondary"):
                    search_text("next")
            
            with search_col3:
                if st.button("✕ Limpiar", key="btn_clear", use_container_width=True,
                            help="Limpiar resultados de búsqueda",
                            type="secondary"):
                    st.session_state.search_term = ""
                    st.session_state.search_results = []
                    st.session_state.current_search_index = 0
                    st.rerun()
        
        # Columna 2: Temas
        with col2:
            st.selectbox("Temas", st.session_state.temas, key="selected_theme", index=0)
            
            # Botón Insertar Tema
            st.markdown('<div class="spacer"></div>', unsafe_allow_html=True)
            if st.button("📝 Insertar Tema", key="btn_insert_theme", use_container_width=True,
                        help="Insertar el tema seleccionado en el editor",
                        type="primary"):
                insert_theme()
        
        # Columna 3: Comentarios
        with col3:
            if st.button("➕ Comentario", key="btn_comment", use_container_width=True, 
                        help="Agregar nuevo comentario", 
                        type="secondary"):
                pass  # Aquí iría la lógica para agregar comentarios
                
            st.markdown('<div class="spacer"></div>', unsafe_allow_html=True)
            if st.button("🔍 Buscar Comentarios", key="btn_search_comments", use_container_width=True,
                        help="Buscar en comentarios existentes",
                        type="secondary"):
                pass  # Aquí iría la lógica para buscar comentarios
                
            st.markdown('<div class="spacer"></div>', unsafe_allow_html=True)
            if st.button("💾 Guardar", key="btn_save", use_container_width=True,
                        help="Guardar cambios en el documento",
                        type="primary"):
                # Guardar el contenido actual
                st.session_state.db_manager.guardar_hoja(
                    st.session_state.current_tomo, 
                    st.session_state.contenido_tomo)
                st.success("Documento guardado correctamente")
        
        st.markdown('</div>', unsafe_allow_html=True)  # Cierre de toolbar-container
    
    # Subárea inferior (editor) con separación de 5px
    with st.container():
        st.markdown('<div class="editor-container">', unsafe_allow_html=True)
        new_content = st.text_area(
            "Contenido:", 
            st.session_state.contenido_tomo, 
            height=500,  # La altura se controla con CSS
            key="editor_area",
            label_visibility="collapsed"
        )
        st.markdown('</div>', unsafe_allow_html=True)
        
        if new_content != st.session_state.contenido_tomo:
            st.session_state.contenido_tomo = new_content
            st.session_state.db_manager.guardar_hoja(
                st.session_state.current_tomo, 
                st.session_state.contenido_tomo)
    
    st.markdown('</div>', unsafe_allow_html=True)  # Cierre del main-container

if __name__ == "__main__":
    main()



